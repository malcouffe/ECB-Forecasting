#!/usr/bin/env python3
"""Run Moirai-2 recursive forecasts and export QoQ quantiles to the TeamXX template."""

from __future__ import annotations

import argparse
from datetime import datetime
from pathlib import Path
import sys
from typing import Sequence

import numpy as np
import pandas as pd

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from ecb_forecast.compare import (  # noqa: E402
    Moirai2Runner,
    parse_quantiles,
    parse_torch_dtype,
    resolve_device,
)
from ecb_forecast.datasets import (  # noqa: E402
    load_quarterly_dataset,
    prepare_country_context,
    prepare_moirai_inputs,
)
from ecb_forecast.timeseries import quantile_column_name  # noqa: E402

DEFAULT_QUANTILES = ",".join(f"{q:.1f}" for q in np.arange(0.1, 1.0, 0.1))
REQUIRED_QUANTILES = [round(q, 1) for q in np.arange(0.1, 1.0, 0.1)]
TEMPLATE_COLUMNS = [
    "cutoff",
    "oos_date",
    "y_true",
    "y_pred",
    "quantile_0.1",
    "quantile_0.2",
    "quantile_0.3",
    "quantile_0.4",
    "quantile_0.5",
    "quantile_0.6",
    "quantile_0.7",
    "quantile_0.8",
    "quantile_0.9",
]


def parse_args(argv: Sequence[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Produce recursive Moirai-2 forecasts and export them in the TeamXX Excel template.",
    )
    parser.add_argument("--country", type=str, default="DE", help="Country identifier to forecast.")
    parser.add_argument(
        "--dataset",
        type=Path,
        default=None,
        help="Path to the quarterly CSV (defaults to data/processed/investment_<country>_quarterly.csv "
        "and falls back to data/processed/investment_quarterly.csv).",
    )
    parser.add_argument(
        "--target-column",
        type=str,
        default="investment",
        help="Target column containing the level series (default: investment).",
    )
    parser.add_argument(
        "--actual-qoq-column",
        type=str,
        default=None,
        help="Optional column containing the QoQ ground truth. "
        "Defaults to '<target-column>_qoq' when available.",
    )
    parser.add_argument(
        "--output-xlsx",
        type=Path,
        default=None,
        help="Excel workbook that will store the formatted sheet "
        "(default: results/TeamXX_recursive_moirai_<country>_<timestamp>.xlsx).",
    )
    parser.add_argument(
        "--horizon",
        type=int,
        default=4,
        help="Recursive prediction horizon per origin (default: 4 quarters).",
    )
    parser.add_argument(
        "--start-date",
        type=str,
        default="2020-10-01",
        help="First cutoff (inclusive) to launch recursive forecasts (default: 2020-10-01).",
    )
    parser.add_argument(
        "--end-date",
        type=str,
        default="2025-01-01",
        help="Last cutoff (inclusive) to launch recursive forecasts (default: 2025-01-01).",
    )
    parser.add_argument(
        "--quantiles",
        type=str,
        default=DEFAULT_QUANTILES,
        help="Comma-separated list of quantiles requested from Moirai-2 "
        "(default: 0.1 → 0.9 in 0.1 increments).",
    )
    parser.add_argument(
        "--moirai-model",
        type=str,
        default="Salesforce/moirai-2.0-R-small",
        help="Moirai-2 model name on the Hugging Face Hub.",
    )
    parser.add_argument(
        "--device",
        type=str,
        default="auto",
        help="Device to execute Moirai-2 on (default: auto).",
    )
    parser.add_argument(
        "--torch-dtype",
        type=str,
        default="bfloat16",
        help="Torch dtype for Moirai-2 weights (default: bfloat16).",
    )
    parser.add_argument(
        "--context-length",
        type=int,
        default=1680,
        help="Maximum number of past observations sent to Moirai-2 (default: 1680).",
    )
    parser.add_argument(
        "--batch-size",
        type=int,
        default=1,
        help="Batch size used by the Moirai-2 predictor (default: 1).",
    )
    parser.add_argument(
        "--id-column",
        type=str,
        default="country",
        help="Identifier column in the dataset (default: country).",
    )
    parser.add_argument(
        "--timestamp-column",
        type=str,
        default="timestamp",
        help="Timestamp column name in the dataset (default: timestamp).",
    )
    return parser.parse_args(argv)


def resolve_dataset_path(args: argparse.Namespace) -> Path:
    if args.dataset:
        return args.dataset
    per_country = PROJECT_ROOT / "data" / "processed" / f"investment_{args.country}_quarterly.csv"
    if per_country.exists():
        return per_country
    combined = PROJECT_ROOT / "data" / "processed" / "investment_quarterly.csv"
    if combined.exists():
        return combined
    raise FileNotFoundError(
        "Could not find a quarterly dataset. Provide --dataset explicitly or place the processed CSV "
        "under data/processed."
    )


def load_country_frame(
    csv_path: Path,
    country: str,
    id_column: str,
    timestamp_column: str,
) -> pd.DataFrame:
    df, _ = load_quarterly_dataset(
        csv_path,
        id_column=id_column,
        timestamp_column=timestamp_column,
    )
    subset = df[df[id_column] == country].copy()
    if subset.empty:
        available = sorted(df[id_column].astype(str).unique())
        raise ValueError(f"No rows for country '{country}' in {csv_path}. Available ids: {available}")
    subset[timestamp_column] = pd.to_datetime(subset[timestamp_column])
    subset = subset.sort_values(timestamp_column).reset_index(drop=True)
    return subset


def clamp_origin_range(
    subset: pd.DataFrame,
    start_date: pd.Timestamp,
    end_date: pd.Timestamp,
    timestamp_column: str,
) -> tuple[pd.Timestamp, pd.Timestamp]:
    data_start = subset[timestamp_column].min()
    data_end = subset[timestamp_column].max()
    start = max(data_start, start_date)
    end = min(data_end, end_date)
    if start > end:
        raise ValueError(
            f"Date range [{start_date.date()}, {end_date.date()}] lies outside "
            f"the available data [{data_start.date()} → {data_end.date()}]."
        )
    return start, end


def run_recursive_forecasts(
    subset: pd.DataFrame,
    runner: Moirai2Runner,
    country: str,
    start_date: pd.Timestamp,
    end_date: pd.Timestamp,
    horizon: int,
    quantile_levels: Sequence[float],
    id_column: str,
    timestamp_column: str,
    target_column: str,
) -> pd.DataFrame:
    work = subset.dropna(subset=[target_column]).copy()
    origins = sorted(ts for ts in work[timestamp_column].unique() if start_date <= ts <= end_date)
    if not origins:
        raise ValueError("No forecast origins fall inside the requested window.")

    print(
        f"\n🔄 Moirai-2 recursive forecasting for '{country}' ({len(origins)} origins, horizon={horizon}, "
        f"{len(quantile_levels)} quantiles)"
    )
    results: list[pd.DataFrame] = []
    for idx, origin in enumerate(origins, start=1):
        context = subset[subset[timestamp_column] <= origin].copy()
        context = context.dropna(subset=[target_column])
        if len(context) < max(4, horizon):
            print(f"  Skipping origin {origin.date()} (insufficient history).")
            continue
        forecast_context = prepare_country_context(
            context,
            prediction_length=horizon,
            series_id=country,
            id_column=id_column,
            timestamp_column=timestamp_column,
            target_column=target_column,
        )
        values, start_timestamp, freq = prepare_moirai_inputs(
            forecast_context.dataframe,
            timestamp_column=timestamp_column,
            target_column=target_column,
        )
        result = runner.run(
            values=values,
            prediction_length=horizon,
            quantile_levels=quantile_levels,
            forecast_index=forecast_context.forecast_index,
            start_timestamp=start_timestamp,
            freq=freq,
        )
        frame = result.df.copy()
        frame["origin_date"] = origin
        frame["horizon_step"] = range(1, horizon + 1)
        frame["model"] = "Moirai-2"
        results.append(frame)

        if idx % 5 == 0 or idx == len(origins):
            print(f"  Processed {idx}/{len(origins)} origins...")

    if not results:
        raise RuntimeError("Moirai-2 did not produce any forecast rows.")
    combined = pd.concat(results, ignore_index=True)
    combined[timestamp_column] = pd.to_datetime(combined[timestamp_column])
    combined["origin_date"] = pd.to_datetime(combined["origin_date"])
    return combined


def lookup_baseline_level(
    subset: pd.DataFrame,
    timestamp_column: str,
    target_column: str,
    origin: pd.Timestamp,
) -> float | None:
    history = subset[
        (subset[timestamp_column] <= origin) & (subset[target_column].notna())
    ]
    if history.empty:
        return None
    return float(history.iloc[-1][target_column])


def build_actual_qoq_map(
    subset: pd.DataFrame,
    timestamp_column: str,
    target_column: str,
    qoq_column: str | None,
) -> dict[pd.Timestamp, float]:
    subset = subset.sort_values(timestamp_column).reset_index(drop=True)
    timestamps = pd.to_datetime(subset[timestamp_column])
    if qoq_column and qoq_column in subset.columns:
        series = subset[[timestamp_column, qoq_column]].dropna()
        return {
            pd.Timestamp(ts): float(val)
            for ts, val in zip(series[timestamp_column], series[qoq_column], strict=True)
        }
    values = subset[target_column].astype(float)
    qoq = values.pct_change() * 100.0
    valid = qoq.notna()
    return {
        pd.Timestamp(ts): float(val)
        for ts, val in zip(timestamps[valid], qoq[valid], strict=True)
    }


def compute_qoq_percentage(current: float | None, previous: float | None) -> float | None:
    if current is None or previous is None:
        return None
    if pd.isna(current) or pd.isna(previous) or previous == 0:
        return None
    return ((current - previous) / previous) * 100.0


def ensure_required_quantiles(quantiles: Sequence[float]) -> None:
    missing = [
        q
        for q in REQUIRED_QUANTILES
        if not any(np.isclose(q, candidate) for candidate in quantiles)
    ]
    if missing:
        raise ValueError(
            "The template export requires quantiles 0.1 → 0.9 (0.1 increments). "
            f"Missing levels: {', '.join(str(q) for q in missing)}. "
            "Provide them via --quantiles."
        )


def format_forecasts_for_template(
    forecasts: pd.DataFrame,
    subset: pd.DataFrame,
    quantile_levels: Sequence[float],
    timestamp_column: str,
    target_column: str,
    actual_qoq_column: str | None,
) -> pd.DataFrame:
    ensure_required_quantiles(quantile_levels)
    quantile_columns = {
        q: quantile_column_name("moirai", q) for q in REQUIRED_QUANTILES
    }
    missing_columns = [col for col in quantile_columns.values() if col not in forecasts.columns]
    if missing_columns:
        raise ValueError(
            f"Forecast results are missing quantile columns: {', '.join(missing_columns)}. "
            "Ensure Moirai was asked to generate those levels."
        )
    median_quantile = 0.5
    actual_qoq_map = build_actual_qoq_map(
        subset,
        timestamp_column=timestamp_column,
        target_column=target_column,
        qoq_column=actual_qoq_column,
    )

    template_rows: list[dict[str, object]] = []
    grouped = forecasts.sort_values(["origin_date", "horizon_step", timestamp_column]).groupby("origin_date")

    for origin, group in grouped:
        baseline = lookup_baseline_level(subset, timestamp_column, target_column, origin)
        if baseline is None:
            print(f"  Warning: missing baseline level for origin {origin.date()}, skipping those horizons.")
            continue
        prev_levels = {q: baseline for q in REQUIRED_QUANTILES}
        for _, row in group.sort_values("horizon_step").iterrows():
            ts = pd.Timestamp(row[timestamp_column])
            qoq_values: dict[float, float | None] = {}
            for q, column in quantile_columns.items():
                raw_value = row[column]
                current = None if pd.isna(raw_value) else float(raw_value)
                prev_value = prev_levels.get(q)
                qoq_values[q] = compute_qoq_percentage(current, prev_value)
                if current is not None:
                    prev_levels[q] = current
            template_rows.append(
                {
                    "cutoff": pd.Timestamp(origin),
                    "oos_date": ts,
                    "y_true": actual_qoq_map.get(ts),
                    "y_pred": qoq_values.get(median_quantile),
                    "quantile_0.1": qoq_values.get(0.1),
                    "quantile_0.2": qoq_values.get(0.2),
                    "quantile_0.3": qoq_values.get(0.3),
                    "quantile_0.4": qoq_values.get(0.4),
                    "quantile_0.5": qoq_values.get(0.5),
                    "quantile_0.6": qoq_values.get(0.6),
                    "quantile_0.7": qoq_values.get(0.7),
                    "quantile_0.8": qoq_values.get(0.8),
                    "quantile_0.9": qoq_values.get(0.9),
                }
            )

    output = pd.DataFrame(template_rows)
    if output.empty:
        raise RuntimeError("No template rows were generated. Check the warnings above.")
    return output[TEMPLATE_COLUMNS]


def write_template(
    frame: pd.DataFrame,
    output_path: Path,
    sheet_name: str,
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    sheet = sheet_name[:31]
    with pd.ExcelWriter(
        output_path,
        engine="openpyxl",
        date_format="YYYY-MM-DD",
        datetime_format="YYYY-MM-DD",
    ) as writer:
        frame.to_excel(writer, sheet_name=sheet, index=False)

    try:
        from openpyxl import load_workbook
    except ImportError:
        print("⚠️  openpyxl is not installed; skipping Excel date formatting.")
        return

    wb = load_workbook(output_path)
    ws = wb[sheet]
    for row_idx in range(2, ws.max_row + 1):
        for col_idx in (1, 2):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is not None:
                cell.number_format = "YYYY-MM-DD"
    wb.save(output_path)


def resolve_actual_column(args: argparse.Namespace) -> str | None:
    if args.actual_qoq_column:
        return args.actual_qoq_column
    candidate = f"{args.target_column}_qoq"
    return candidate


def main(argv: Sequence[str] | None = None) -> int:
    args = parse_args(argv)
    dataset_path = resolve_dataset_path(args)
    subset = load_country_frame(dataset_path, args.country, args.id_column, args.timestamp_column)

    for column in (args.target_column,):
        if column not in subset.columns:
            raise ValueError(f"Column '{column}' not found in {dataset_path}.")

    start_date = pd.to_datetime(args.start_date)
    end_date = pd.to_datetime(args.end_date)
    start_date, end_date = clamp_origin_range(subset, start_date, end_date, args.timestamp_column)

    quantiles = parse_quantiles(args.quantiles)
    ensure_required_quantiles(quantiles)
    dtype = parse_torch_dtype(args.torch_dtype)
    device = resolve_device(args.device)
    if device == "mps":
        print("⚠️  MPS detected for Moirai-2; falling back to CPU to avoid float64 limitations.")
        device = "cpu"
    runner = Moirai2Runner(
        model_name=args.moirai_model,
        device=device,
        dtype=dtype,
        context_length=args.context_length,
        batch_size=args.batch_size,
    )

    forecasts = run_recursive_forecasts(
        subset=subset,
        runner=runner,
        country=args.country,
        start_date=start_date,
        end_date=end_date,
        horizon=args.horizon,
        quantile_levels=quantiles,
        id_column=args.id_column,
        timestamp_column=args.timestamp_column,
        target_column=args.target_column,
    )

    actual_column = resolve_actual_column(args)
    template = format_forecasts_for_template(
        forecasts=forecasts,
        subset=subset,
        quantile_levels=quantiles,
        timestamp_column=args.timestamp_column,
        target_column=args.target_column,
        actual_qoq_column=actual_column,
    )
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    if args.output_xlsx is None:
        output_path = PROJECT_ROOT / "results" / f"TeamXX_recursive_moirai_{args.country}_{timestamp}.xlsx"
    else:
        output_path = args.output_xlsx
    sheet_name = f"{args.target_column}_{args.country}"
    write_template(template, output_path, sheet_name)

    print("\n✅ QoQ export ready!")
    print(f"  Template Excel: {output_path}")
    print(f"  Sheet: {sheet_name[:31]}")
    print(f"  Rows: {len(template)} covering {template['oos_date'].min()} → {template['oos_date'].max()}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
