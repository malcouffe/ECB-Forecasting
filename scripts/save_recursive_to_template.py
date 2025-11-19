#!/usr/bin/env python3
"""Save recursive forecasts to the TeamXX template Excel file."""

import pandas as pd
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parents[1]
RESULTS_DIR = PROJECT_ROOT / "results"
TEMPLATE_PATH = PROJECT_ROOT / "TeamXX_template_investment.xlsx"

def load_recursive_forecasts():
    """Load both Chronos and Moirai recursive forecasts."""
    recursive_dir = RESULTS_DIR / "recursive"
    chronos_df = pd.read_csv(recursive_dir / "recursive_chronos_DE_4q.csv")
    moirai_df = pd.read_csv(recursive_dir / "recursive_moirai_DE_4q.csv")
    
    chronos_df['timestamp'] = pd.to_datetime(chronos_df['timestamp'])
    moirai_df['timestamp'] = pd.to_datetime(moirai_df['timestamp'])
    
    return chronos_df, moirai_df

def interpolate_quantiles(p10, p50, p90):
    """
    Interpolate missing quantiles (0.2, 0.3, 0.4, 0.6, 0.7, 0.8) from available quantiles.
    Uses linear interpolation between known quantiles.
    """
    # Known quantiles: 0.1, 0.5, 0.9
    # We need: 0.2, 0.3, 0.4, 0.6, 0.7, 0.8
    
    # Between 0.1 and 0.5 (span of 0.4)
    q02 = p10 + (p50 - p10) * (0.1 / 0.4)  # 0.2 is 0.1 above 0.1
    q03 = p10 + (p50 - p10) * (0.2 / 0.4)  # 0.3 is 0.2 above 0.1
    q04 = p10 + (p50 - p10) * (0.3 / 0.4)  # 0.4 is 0.3 above 0.1
    
    # Between 0.5 and 0.9 (span of 0.4)
    q06 = p50 + (p90 - p50) * (0.1 / 0.4)  # 0.6 is 0.1 above 0.5
    q07 = p50 + (p90 - p50) * (0.2 / 0.4)  # 0.7 is 0.2 above 0.5
    q08 = p50 + (p90 - p50) * (0.3 / 0.4)  # 0.8 is 0.3 above 0.5
    
    return q02, q03, q04, q06, q07, q08

def get_previous_value(moirai_df, current_idx, current_row):
    """
    Get the previous quarter's value for QoQ calculation.
    
    For the first forecast from a cutoff (horizon_step=1), use the actual value
    from the cutoff date. For subsequent forecasts, use the predicted value
    from the previous horizon.
    """
    if current_row['horizon_step'] == 1:
        # First step: use actual value at cutoff date (origin_date)
        # This should be the last known actual value
        cutoff_data = moirai_df[
            (moirai_df['timestamp'] == current_row['origin_date']) &
            (moirai_df['actual'].notna())
        ]
        if len(cutoff_data) > 0:
            return cutoff_data.iloc[0]['actual']
        else:
            # If no actual at cutoff, find the most recent actual before cutoff
            earlier_data = moirai_df[
                (moirai_df['timestamp'] < current_row['origin_date']) &
                (moirai_df['actual'].notna())
            ].sort_values('timestamp', ascending=False)
            if len(earlier_data) > 0:
                return earlier_data.iloc[0]['actual']
    else:
        # Subsequent steps: use predicted value from previous step
        prev_step = moirai_df[
            (moirai_df['origin_date'] == current_row['origin_date']) &
            (moirai_df['horizon_step'] == current_row['horizon_step'] - 1)
        ]
        if len(prev_step) > 0:
            return prev_step.iloc[0]['moirai_p50']
    
    # Fallback: return None if we can't find previous value
    return None

def compute_qoq_percentage(current_value, previous_value):
    """Compute quarter-on-quarter percentage change."""
    if previous_value is None or previous_value == 0 or pd.isna(previous_value):
        return None
    return ((current_value - previous_value) / previous_value) * 100

def save_to_template(
    chronos_df,
    moirai_df,
    country='DE',
    output_path: str | Path | None = None,
):
    """
    Save all recursive forecasts to the Excel template format.
    
    The template expects:
    - cutoff: the date up to which data was used (origin_date) 
    - oos_date: the forecast date (timestamp)
    - y_true: actual QoQ percentage change if available
    - y_pred: predicted QoQ percentage change (median)
    - quantile_0.1 through quantile_0.9: quantile predictions (QoQ %)
    """
    output_path = Path(output_path) if output_path else RESULTS_DIR / "TeamXX_recursive_investment_predictions.xlsx"
    
    # Convert date columns to datetime
    moirai_df = moirai_df.copy()
    moirai_df['origin_date'] = pd.to_datetime(moirai_df['origin_date'])
    moirai_df['timestamp'] = pd.to_datetime(moirai_df['timestamp'])
    
    # Sort by origin_date and timestamp
    moirai_df = moirai_df.sort_values(['origin_date', 'timestamp']).reset_index(drop=True)
    
    # Load the quarterly data to get actual QoQ values
    quarterly_file = PROJECT_ROOT / "data" / "processed" / f"investment_{country}_quarterly.csv"
    quarterly_df = pd.read_csv(quarterly_file)
    quarterly_df['timestamp'] = pd.to_datetime(quarterly_df['timestamp'])
    
    # Prepare the data in the template format
    forecast_data = []
    
    for idx, row in moirai_df.iterrows():
        # Get previous quarter's value for QoQ calculation
        prev_value = get_previous_value(moirai_df, idx, row)
        
        if prev_value is None:
            print(f"Warning: Could not find previous value for {row['timestamp']} (origin: {row['origin_date']})")
            continue
        
        # Compute QoQ percentages for all quantiles
        qoq_p10 = compute_qoq_percentage(row['moirai_p10'], prev_value)
        qoq_p50 = compute_qoq_percentage(row['moirai_p50'], prev_value)
        qoq_p90 = compute_qoq_percentage(row['moirai_p90'], prev_value)
        
        # Interpolate missing quantiles (using QoQ percentages)
        q02, q03, q04, q06, q07, q08 = interpolate_quantiles(
            qoq_p10, 
            qoq_p50, 
            qoq_p90
        )
        
        # Get actual QoQ from quarterly data if available
        actual_qoq = None
        actual_data = quarterly_df[quarterly_df['timestamp'] == row['timestamp']]
        if len(actual_data) > 0 and pd.notna(actual_data.iloc[0]['investment_qoq']):
            actual_qoq = actual_data.iloc[0]['investment_qoq']
        
        forecast_row = {
            'cutoff': row['origin_date'],  # Keep as pandas Timestamp for proper Excel formatting
            'oos_date': row['timestamp'],  # Keep as pandas Timestamp for proper Excel formatting
            'y_true': actual_qoq,
            'y_pred': qoq_p50,  # Use median QoQ % as prediction
            'quantile_0.1': qoq_p10,
            'quantile_0.2': q02,
            'quantile_0.3': q03,
            'quantile_0.4': q04,
            'quantile_0.5': qoq_p50,
            'quantile_0.6': q06,
            'quantile_0.7': q07,
            'quantile_0.8': q08,
            'quantile_0.9': qoq_p90
        }
        forecast_data.append(forecast_row)
    
    # Create the output dataframe
    output_df = pd.DataFrame(forecast_data)
    
    # Write to Excel with the country-specific sheet name and proper date formatting
    sheet_name = f'investment_{country}'
    with pd.ExcelWriter(output_path, engine='openpyxl', date_format='YYYY-MM-DD', datetime_format='YYYY-MM-DD') as writer:
        output_df.to_excel(writer, sheet_name=sheet_name, index=False)
    
    # Post-process to ensure dates are formatted correctly in Excel
    from openpyxl import load_workbook
    wb = load_workbook(output_path)
    ws = wb[sheet_name]
    
    # Format the first two columns (cutoff and oos_date) as dates
    for row_idx in range(2, ws.max_row + 1):  # Start from row 2 (skip header)
        for col_idx in [1, 2]:  # Columns A and B (cutoff and oos_date)
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is not None:
                cell.number_format = 'YYYY-MM-DD'
    
    wb.save(output_path)
    
    print(f"✅ Recursive forecasts (QoQ %) saved to: {output_path}")
    print(f"   Sheet name: {sheet_name}")
    print(f"   Total forecasts: {len(output_df)}")
    print(f"   Date range: {output_df['oos_date'].min()} to {output_df['oos_date'].max()}")
    print(f"   Cutoff dates: {output_df['cutoff'].nunique()} unique origins")
    print(f"   Format: Quarter-on-Quarter percentage changes")
    
    return output_path

def main():
    """Main execution."""
    print("=" * 80)
    print("📊 Saving Recursive Forecasts to Excel Template (QoQ %)")
    print("=" * 80)
    
    print("\n1. Loading recursive forecasts...")
    chronos_df, moirai_df = load_recursive_forecasts()
    
    print(f"   ✓ Chronos forecasts: {len(chronos_df)} rows")
    print(f"   ✓ Moirai forecasts: {len(moirai_df)} rows")
    
    print(f"\n2. Forecast details:")
    print(f"   Origin dates: {moirai_df['origin_date'].nunique()} unique cutoff points")
    print(f"   Date range: {moirai_df['timestamp'].min()} to {moirai_df['timestamp'].max()}")
    print(f"   Horizons: {sorted(moirai_df['horizon_step'].unique())}")
    
    print("\n3. Converting absolute predictions to QoQ percentages...")
    print("   (Computing quarter-on-quarter percentage changes)")
    
    print("\n4. Saving to Excel template format...")
    output_path = save_to_template(chronos_df, moirai_df, country='DE')
    
    print(f"\n{'=' * 80}")
    print(f"✅ Done! Output saved to: {output_path}")
    print(f"   All predictions are now in Quarter-on-Quarter % format")
    print(f"{'=' * 80}")

if __name__ == '__main__':
    main()
